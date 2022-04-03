from tkinter import * 
import random
import tkinter.messagebox as  tkMessageBox
from openpyxl import load_workbook
workbook = load_workbook(filename="project.xlsx")
workbook.sheetnames
sheet = workbook.active
List1=['A','B','C','D','E']
A=List1[0]
B=List1[1]
C=List1[2]
D=List1[3]
E=List1[4]
a=random.randint(1,15)
Ation=sheet['A%d'%a].value
Rmance=sheet['B%d'%a].value
Sorts =sheet['C%d'%a].value
Cmedy =sheet['D%d'%a].value
Hrror =sheet['E%d'%a].value
top = Tk()
top.title("Random Anime")
top.geometry("400x400")
def helloCallBack1():
   tkMessageBox.showinfo( "Go watch",Ation)
def helloCallBack2():
   tkMessageBox.showinfo( "Go watch",Rmance)
def helloCallBack3():
   tkMessageBox.showinfo( "Go watch",Sorts)
def helloCallBack4():
   tkMessageBox.showinfo( "Go watch",Cmedy)
def helloCallBack5():
   tkMessageBox.showinfo( "Go watch",Hrror)
Action= Button(top, text = "Action",activebackground = "black", activeforeground = "white",command = helloCallBack1).place(x = 20, y = 25)
Romance= Button(top, text = "Romance",activebackground = "black", activeforeground = "white",command = helloCallBack2).place(x = 90, y = 25)
Sports = Button(top, text = "Sports",activebackground = "black", activeforeground = "white",command = helloCallBack3).place(x = 175, y = 25)
Comedy = Button(top, text = "Comedy",activebackground = "black", activeforeground = "white",command = helloCallBack4).place(x = 245, y = 25)
Horror = Button(top, text = "Horror",activebackground = "black", activeforeground = "white",command = helloCallBack5).place(x = 325, y = 25)
top.mainloop()